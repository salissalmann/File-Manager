"""Parse ledger Excel files into structured LedgerRow objects."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl

from src.models import LedgerRow


def _detect_format(ws) -> str:
    """Detect ledger format from header row. Returns 'new' or 'legacy'."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_strs = [str(h).strip().lower() if h else "" for h in headers]
    # New format: has "Account" at col 6 with embedded CA codes
    if len(header_strs) > 6 and header_strs[6] == "account":
        return "new"
    return "legacy"


def parse_ledger(file_path: str | Path) -> list[LedgerRow]:
    """Read a ledger Excel file and return a list of LedgerRow objects.

    Auto-detects format:
      - Legacy: Vendor, Date, CA Account, Folder, Amount
      - New:    Vendor, Date, Num, Column2, Memo, Column3, Account, ..., Amount
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    fmt = _detect_format(ws)
    rows: list[LedgerRow] = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not row or row[0] is None:
            continue

        vendor = _normalize_vendor(row[0])
        dt = _parse_date(row[1])

        if fmt == "new":
            ca_code = _extract_ca_from_account(row[6]) if len(row) > 6 else ""
            folder = None
            amount = _parse_amount(row[9]) if len(row) > 9 else None
        else:
            ca_code = _normalize_ca(row[2])
            folder = str(row[3]).strip() if row[3] else None
            amount = _parse_amount(row[4])

        if vendor and dt and amount is not None:
            rows.append(LedgerRow(
                index=idx - 1,  # 1-based data row index
                vendor=vendor,
                date=dt,
                ca_code=ca_code,
                folder=folder,
                amount=amount,
            ))

    wb.close()
    return rows


def _normalize_vendor(value: object) -> str:
    """Normalize vendor name: strip whitespace, preserve original case."""
    return str(value).strip()


def _parse_date(value: object) -> date | None:
    """Parse a date from Excel cell (may be datetime or string)."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _extract_ca_from_account(value: object) -> str:
    """Extract CA code from Account column, e.g. '50500 · Equipment Rental' → '50500'."""
    if value is None:
        return ""
    text = str(value).strip()
    # Account format: "50500 · Description" — split on middot separator
    if " · " in text:
        return text.split(" · ", 1)[0].strip()
    # Fallback: take leading digits
    digits = ""
    for ch in text:
        if ch.isdigit():
            digits += ch
        elif digits:
            break
    return digits


def _normalize_ca(value: object) -> str:
    """Normalize CA code: convert to string, strip whitespace."""
    if value is None:
        return ""
    # Handle numeric CA codes (Excel may store as int/float)
    if isinstance(value, (int, float)):
        return str(int(value))
    return str(value).strip()


def _parse_amount(value: object) -> Decimal | None:
    """Parse amount from Excel cell."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        cleaned = value.strip().replace("$", "").replace(",", "")
        try:
            return Decimal(cleaned)
        except Exception:
            return None
    return None
