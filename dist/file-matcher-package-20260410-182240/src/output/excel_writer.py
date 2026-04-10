"""Generate clean Excel output with match results, orphans, and veto log."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import config
from src.engine.explainer import explain, explain_ambiguity, format_reasons, format_signals
from src.models import FileRecord, MatchResult
from src.output.sharepoint_links import resolve_file_open_url


# 5-tier status colors
_FILLS = {
    config.TIER_CONFIDENT: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # green
    config.TIER_PROBABLE: PatternFill(start_color="B4D6A4", end_color="B4D6A4", fill_type="solid"),   # light green
    config.TIER_POSSIBLE: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),   # yellow
    config.TIER_REVIEW: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),     # salmon
    config.TIER_NO_MATCH: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),   # gray
}

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_LINK_FONT = Font(color="0563C1", underline="single")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def write_results(
    results: list[MatchResult],
    orphans: list[FileRecord],
    output_path: str | Path,
    base_path: str = "",
    path_by_filename: dict[str, str] | None = None,
) -> None:
    """Write matching results to an Excel file with multiple sheets.

    path_by_filename: maps matched filename → inventory path (e.g. CSV "Path From Root").
    When set and base_path is empty, File Path uses SharePoint-style URLs from sharepoint_links.
    """
    wb = Workbook()

    _write_matches_sheet(wb.active, results, base_path, path_by_filename)
    _write_orphans_sheet(wb.create_sheet("Orphan Files"), orphans)
    _write_veto_log_sheet(wb.create_sheet("Veto Log"), results)
    _write_summary_sheet(wb.create_sheet("Summary"), results, orphans)

    # Move Summary sheet to first position
    wb.move_sheet("Summary", offset=-3)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _file_path_cell(
    filename: str,
    base_path: str,
    path_by_filename: dict[str, str] | None,
) -> tuple[str, str | None]:
    """(cell value, hyperlink target or None)."""
    inv_path = ""
    if path_by_filename is not None:
        inv_path = path_by_filename.get(filename, "").strip()
    if base_path:
        rel = inv_path or filename
        rel = rel.replace("\\", "/")
        base = base_path.rstrip("/").rstrip("\\")
        url = f"{base}/{rel}"
        return url, url
    if path_by_filename is not None:
        url = resolve_file_open_url(inv_path, filename)
        return (url or filename), (url or None)
    return filename, None


_TIER_ORDER = {"Confident": 0, "Probable": 1, "Possible": 2, "Review": 3, "No Match": 4}


def _write_matches_sheet(
    ws,
    results: list[MatchResult],
    base_path: str,
    path_by_filename: dict[str, str] | None,
) -> None:
    """Write the main matches sheet, sorted by status tier then score descending."""
    ws.title = "Match Results"
    results = sorted(results, key=lambda r: (_TIER_ORDER.get(r.status, 9), -r.score))

    headers = [
        "Row #", "Vendor", "Date", "CA Code", "Amount",
        "Matched File", "File Path", "Score", "Signals",
        "Status", "Explanation", "Decision Trail",
        "Veto Info", "Alternatives", "Ambiguity Flag",
    ]
    _write_header(ws, headers)

    for i, r in enumerate(results, start=2):
        row = r.ledger_row
        ws.cell(row=i, column=1, value=row.index)
        ws.cell(row=i, column=2, value=row.vendor)
        ws.cell(row=i, column=3, value=str(row.date))
        ws.cell(row=i, column=4, value=row.ca_code)
        ws.cell(row=i, column=5, value=float(row.amount))

        # Matched file name
        file_name = r.matched_file.original_name if r.matched_file else "—"
        ws.cell(row=i, column=6, value=file_name)

        # File path with clickable hyperlink when base_path or path_by_filename is set
        if r.matched_file:
            display, target = _file_path_cell(
                r.matched_file.original_name, base_path, path_by_filename
            )
            path_cell = ws.cell(row=i, column=7, value=display)
            if target:
                path_cell.hyperlink = target
                path_cell.font = _LINK_FONT
        else:
            ws.cell(row=i, column=7, value="—")

        # Score + Signals + Status
        ws.cell(row=i, column=8, value=round(r.score, 1))
        ws.cell(row=i, column=9, value=format_signals(r))
        ws.cell(row=i, column=10, value=r.status)

        # Plain-English explanation
        ws.cell(row=i, column=11, value=explain(r))

        # Decision trail (plain-English reasons + ambiguity)
        trail = format_reasons(r.reasons)
        ambiguity_text = explain_ambiguity(r)
        if ambiguity_text:
            trail += "\n" + ambiguity_text
        ws.cell(row=i, column=12, value=trail)
        ws.cell(row=i, column=12).alignment = Alignment(wrap_text=True)

        # Veto info
        veto_info = ""
        if r.vetoed_candidates:
            top_veto = r.vetoed_candidates[0]
            veto_info = f"{top_veto.file.original_name}: {top_veto.veto_reason}"
        ws.cell(row=i, column=13, value=veto_info)

        # Alternatives
        alt_text = ""
        if r.alternatives:
            alt_parts = []
            for alt in r.alternatives[:3]:
                alt_parts.append(
                    f"{alt.file.original_name} (score={alt.score:.0f}, {format_reasons(alt.reasons)})"
                )
            alt_text = " | ".join(alt_parts)
        ws.cell(row=i, column=14, value=alt_text)

        # Ambiguity flag
        ws.cell(row=i, column=15, value="Yes" if r.is_ambiguous else "")

        # Apply status color to the entire row
        fill = _FILLS.get(r.status)
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = fill
                ws.cell(row=i, column=col).border = _THIN_BORDER

        # Wrap the explanation column
        ws.cell(row=i, column=11).alignment = Alignment(wrap_text=True)

    _auto_width(ws, headers)


def _write_orphans_sheet(ws, orphans: list[FileRecord]) -> None:
    """Write orphan files that matched no ledger row."""
    headers = ["File Name", "Parsed Vendor", "Parsed CA Codes", "Parsed Amounts", "Parsed Dates"]
    _write_header(ws, headers)

    for i, fr in enumerate(orphans, start=2):
        ws.cell(row=i, column=1, value=fr.original_name)
        ws.cell(row=i, column=2, value=fr.vendor_name)
        ws.cell(row=i, column=3, value=", ".join(fr.ca_codes) if fr.ca_codes else "—")
        ws.cell(row=i, column=4, value=", ".join(str(a) for a in fr.amounts) if fr.amounts else "—")
        ws.cell(row=i, column=5, value=", ".join(str(d) for d in fr.dates) if fr.dates else "—")
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = _THIN_BORDER

    _auto_width(ws, headers)


def _write_veto_log_sheet(ws, results: list[MatchResult]) -> None:
    """Write detailed veto log for all rejected matches."""
    headers = ["Ledger Row #", "Ledger Vendor", "Ledger CA", "Ledger Amount",
               "Rejected File", "File Score", "Veto Reason"]
    _write_header(ws, headers)

    row_num = 2
    for r in results:
        for vc in r.vetoed_candidates:
            ws.cell(row=row_num, column=1, value=r.ledger_row.index)
            ws.cell(row=row_num, column=2, value=r.ledger_row.vendor)
            ws.cell(row=row_num, column=3, value=r.ledger_row.ca_code)
            ws.cell(row=row_num, column=4, value=float(r.ledger_row.amount))
            ws.cell(row=row_num, column=5, value=vc.file.original_name)
            ws.cell(row=row_num, column=6, value=round(vc.score, 1))
            ws.cell(row=row_num, column=7, value=vc.veto_reason or "")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).border = _THIN_BORDER
            row_num += 1

    _auto_width(ws, headers)


def _write_summary_sheet(ws, results: list[MatchResult], orphans: list[FileRecord]) -> None:
    """Write a summary overview sheet."""
    headers = ["Metric", "Count"]
    _write_header(ws, headers)

    total = len(results)
    confident = sum(1 for r in results if r.status == config.TIER_CONFIDENT)
    probable = sum(1 for r in results if r.status == config.TIER_PROBABLE)
    possible = sum(1 for r in results if r.status == config.TIER_POSSIBLE)
    review = sum(1 for r in results if r.status == config.TIER_REVIEW)
    no_match = sum(1 for r in results if r.status == config.TIER_NO_MATCH)
    ambiguous = sum(1 for r in results if r.is_ambiguous)

    matched = confident + probable
    metrics = [
        ("Total Ledger Rows", total),
        ("Confident Matches", confident),
        ("Probable Matches", probable),
        ("Possible Matches", possible),
        ("Needs Review", review),
        ("No Match", no_match),
        ("Ambiguous Matches", ambiguous),
        ("Orphan Files", len(orphans)),
        ("Match Rate (Confident + Probable)", f"{(matched / total * 100):.1f}%" if total else "0%"),
    ]

    for i, (label, value) in enumerate(metrics, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=1).font = Font(bold=True)
        for col in range(1, 3):
            ws.cell(row=i, column=col).border = _THIN_BORDER

    _auto_width(ws, headers)


def _write_header(ws, headers: list[str]) -> None:
    """Write styled header row."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER


def _auto_width(ws, headers: list[str]) -> None:
    """Auto-adjust column widths based on content."""
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            val = str(row[0]) if row[0] else ""
            max_len = max(max_len, min(len(val), 60))
        ws.column_dimensions[col_letter].width = max_len + 3
