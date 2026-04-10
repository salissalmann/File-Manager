"""Tests for the Excel output writer — file links, alternatives, sheet structure."""

from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.models import FileRecord, LedgerRow, MatchResult, ScoredCandidate
from src.output.excel_writer import write_results


@pytest.fixture
def sample_results(tmp_path):
    """Build a minimal results set and write to Excel."""
    lr = LedgerRow(index=1, vendor="Deluxe", date=date(2021, 11, 19),
                   ca_code="65100", folder=None, amount=Decimal("394.55"))

    matched = FileRecord(original_name="deluxe_checks.pdf", vendor_tokens=["deluxe"],
                         ca_codes=["65100"], amounts=[Decimal("394.55")])

    alt_file = FileRecord(original_name="alt_deluxe.pdf", vendor_tokens=["deluxe"],
                          ca_codes=["65100"], amounts=[Decimal("394.55")])

    alt_candidate = ScoredCandidate(file=alt_file, score=75.0,
                                    reasons=["amount_exact", "ca_match"])

    result = MatchResult(
        ledger_row=lr,
        matched_file=matched,
        score=100.0,
        reasons=["amount_exact", "ca_match", "vendor_match"],
        status="Confident",
        alternatives=[alt_candidate],
    )

    orphan = FileRecord(original_name="20211206_161057")
    output_path = tmp_path / "test_output.xlsx"

    return result, orphan, output_path


class TestExcelSheetStructure:
    def test_four_sheets_created(self, sample_results):
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path)
        wb = load_workbook(output_path)
        assert set(wb.sheetnames) == {"Summary", "Match Results", "Orphan Files", "Veto Log"}

    def test_match_results_headers(self, sample_results):
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 16)]
        assert "File Path" in headers
        assert "Alternatives" in headers
        assert "Signals" in headers
        assert "Explanation" in headers
        assert "Decision Trail" in headers
        assert "Ambiguity Flag" in headers


class TestFilePathColumn:
    def test_file_path_without_base_path(self, sample_results):
        """Without base_path and without path_by_filename, File Path shows filename only."""
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path, base_path="")
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        file_path_val = ws.cell(row=2, column=7).value
        assert file_path_val == "deluxe_checks.pdf"
        assert ws.cell(row=2, column=7).hyperlink is None

    def test_file_path_with_path_map_sharepoint_hyperlink(self, sample_results, monkeypatch):
        """With path_by_filename and no base_path, File Path is a SharePoint URL with hyperlink."""
        monkeypatch.delenv("FILE_LINK_BASE", raising=False)
        monkeypatch.delenv("NEXT_PUBLIC_FILE_LINK_BASE", raising=False)
        result, orphan, output_path = sample_results
        paths = {"deluxe_checks.pdf": "11-2021/deluxe_checks.pdf"}
        write_results([result], [orphan], output_path, base_path="", path_by_filename=paths)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        cell = ws.cell(row=2, column=7)
        assert cell.hyperlink is not None
        target = str(cell.hyperlink.target)
        assert "AllItems.aspx" in target
        assert "viewid=" in target
        assert "parent=" in target

    def test_file_path_with_base_path(self, sample_results):
        """With base_path, File Path column shows full path."""
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path,
                      base_path="https://sharepoint.com/files")
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        file_path_val = ws.cell(row=2, column=7).value
        assert file_path_val == "https://sharepoint.com/files/deluxe_checks.pdf"

    def test_file_path_hyperlink_with_base_path(self, sample_results):
        """With base_path, the cell should have a clickable hyperlink."""
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path,
                      base_path="https://sharepoint.com/files")
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        cell = ws.cell(row=2, column=7)
        assert cell.hyperlink is not None
        assert "sharepoint.com" in str(cell.hyperlink.target)

    def test_no_match_file_path_shows_dash(self, sample_results):
        """No-match row should show '—' in file path column."""
        result, orphan, output_path = sample_results
        no_match = MatchResult(
            ledger_row=result.ledger_row,
            matched_file=None,
            score=0.0,
            status="No Match",
        )
        write_results([no_match], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        assert ws.cell(row=2, column=7).value == "—"


class TestAlternativesColumn:
    def test_alternatives_shown(self, sample_results):
        """Alternatives column should list runner-up candidates (column 14)."""
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        alt_val = ws.cell(row=2, column=14).value
        assert "alt_deluxe.pdf" in alt_val
        assert "score=75" in alt_val

    def test_no_alternatives_empty(self, sample_results):
        """Row with no alternatives should have empty alternatives cell."""
        result, orphan, output_path = sample_results
        result.alternatives = []
        write_results([result], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        alt_val = ws.cell(row=2, column=14).value
        assert alt_val in ("", None)


class TestExplanationColumn:
    def test_explanation_populated(self, sample_results):
        """Explanation column should contain plain-English text."""
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        explanation = ws.cell(row=2, column=11).value
        assert "Confident match" in explanation
        assert "$394" in explanation or "$394.55" in explanation

    def test_signals_column_format(self, sample_results):
        """Signals column shows signal count in 'N/3 (...)' format."""
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        signals_val = ws.cell(row=2, column=9).value
        assert "3/3" in signals_val


class TestAmbiguityFlag:
    def test_not_ambiguous_empty(self, sample_results):
        """Non-ambiguous row has empty ambiguity flag."""
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        flag = ws.cell(row=2, column=15).value
        assert flag in ("", None)

    def test_ambiguous_shows_yes(self, sample_results):
        """Ambiguous row has 'Yes' in ambiguity flag."""
        result, orphan, output_path = sample_results
        result.is_ambiguous = True
        write_results([result], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        flag = ws.cell(row=2, column=15).value
        assert flag == "Yes"


class TestExcelSummaryAndStyle:
    def test_summary_stats(self, sample_results):
        """Summary sheet should have correct counts and match rate format."""
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Summary"]
        # Metric labels are in column 1, values in column 2
        metrics = {ws.cell(row=i, column=1).value: ws.cell(row=i, column=2).value for i in range(2, 11)}
        assert metrics["Confident Matches"] == 1
        assert metrics["Orphan Files"] == 1
        assert "100.0%" in str(metrics["Match Rate (Confident + Probable)"])

    def test_row_color_coding(self, sample_results):
        """Confident matches should be green, No Matches gray."""
        result, orphan, output_path = sample_results
        no_match = MatchResult(ledger_row=result.ledger_row, matched_file=None, score=0, status="No Match")
        write_results([result, no_match], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        # Confident row (green: C6EFCE)
        fill_green = ws.cell(row=2, column=1).fill.start_color.index
        assert fill_green == "00C6EFCE" or fill_green == "C6EFCE"
        # No Match row (gray: D9D9D9)
        fill_gray = ws.cell(row=3, column=1).fill.start_color.index
        assert fill_gray == "00D9D9D9" or fill_gray == "D9D9D9"

    def test_match_results_sort_order(self, sample_results):
        """Result rows should be sorted by tier (Confident first, No Match last)."""
        result_confident, orphan, output_path = sample_results
        result_no_match = MatchResult(
            ledger_row=LedgerRow(index=2, vendor="TACOS CHALES", date=date(2022,1,4), ca_code="53604", folder=None, amount=Decimal("132.61")),
            matched_file=None, score=0, status="No Match"
        )
        # Even if we pass them in reverse order
        write_results([result_no_match, result_confident], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        assert ws.cell(row=2, column=10).value == "Confident"
        assert ws.cell(row=3, column=10).value == "No Match"

    def test_empty_results_no_crash(self, tmp_path):
        """Empty results/orphans should not crash and still create sheets."""
        output_path = tmp_path / "empty.xlsx"
        write_results([], [], output_path)
        assert output_path.exists()
        wb = load_workbook(output_path)
        assert "Summary" in wb.sheetnames

    def test_all_no_match_rate(self, sample_results):
        """All No Match results should show 0.0% match rate."""
        result, orphan, output_path = sample_results
        no_match = MatchResult(ledger_row=result.ledger_row, matched_file=None, score=0, status="No Match")
        write_results([no_match], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Summary"]
        rate = ws.cell(row=10, column=2).value
        assert rate == "0.0%"

    def test_veto_log_sheet_populated(self, sample_results):
        """Vetoed candidates should appear in the Veto Log sheet."""
        result, orphan, output_path = sample_results
        vetoed = ScoredCandidate(file=orphan, score=50, reasons=["ca_match"], vetoed=True, veto_reason="Amount conflict")
        result_no_match = MatchResult(
            ledger_row=result.ledger_row, matched_file=None, score=0, status="No Match",
            vetoed_candidates=[vetoed]
        )
        write_results([result_no_match], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Veto Log"]
        assert ws.cell(row=2, column=5).value == "20211206_161057"
        assert ws.cell(row=2, column=7).value == "Amount conflict"
