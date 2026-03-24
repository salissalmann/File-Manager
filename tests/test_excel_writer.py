"""Tests for the Excel output writer — file links, alternatives, sheet structure."""

from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.models import FileRecord, LedgerRow, MatchResult, ScoredCandidate
from src.output.excel_writer import write_results


@pytest.fixture
def sample_results(tmp_path):
    """Build a minimal results set and write to Excel."""
    lr = LedgerRow(index=1, vendor="Test Vendor", date=date(2023, 1, 1),
                   ca_code="50408", folder=None, amount=Decimal("1000"))

    matched = FileRecord(original_name="test_file.pdf", vendor_tokens=["test"],
                         ca_codes=["50408"], amounts=[Decimal("1000")])

    alt_file = FileRecord(original_name="alt_file.pdf", vendor_tokens=["test"],
                          ca_codes=["50408"], amounts=[Decimal("1000")])

    alt_candidate = ScoredCandidate(file=alt_file, score=75.0,
                                    reasons=["amount_exact", "ca_match"])

    result = MatchResult(
        ledger_row=lr,
        matched_file=matched,
        score=100.0,
        reasons=["amount_exact", "ca_match", "vendor_match"],
        status="Strong",
        alternatives=[alt_candidate],
    )

    orphan = FileRecord(original_name="orphan.pdf")
    output_path = tmp_path / "test_output.xlsx"

    return result, orphan, output_path


class TestExcelSheetStructure:
    def test_four_sheets_created(self, sample_results):
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path)
        wb = load_workbook(output_path)
        assert set(wb.sheetnames) == {"Summary", "Match Results", "Orphan Files", "Veto Log"}

    def test_match_results_headers_include_file_path(self, sample_results):
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 14)]
        assert "File Path" in headers
        assert "Alternatives" in headers


class TestFilePathColumn:
    def test_file_path_without_base_path(self, sample_results):
        """Without base_path, File Path column shows just the filename."""
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path, base_path="")
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        file_path_val = ws.cell(row=2, column=7).value
        assert file_path_val == "test_file.pdf"

    def test_file_path_with_base_path(self, sample_results):
        """With base_path, File Path column shows full path."""
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path,
                      base_path="https://sharepoint.com/files")
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        file_path_val = ws.cell(row=2, column=7).value
        assert file_path_val == "https://sharepoint.com/files/test_file.pdf"

    def test_file_path_hyperlink_with_base_path(self, sample_results):
        """With base_path, the cell should have a clickable hyperlink."""
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path,
                      base_path="https://sharepoint.com/files")
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        cell = ws.cell(row=2, column=7)
        assert cell.hyperlink is not None
        assert "sharepoint.com" in str(cell.hyperlink.target)

    def test_no_match_file_path_shows_dash(self, sample_results):
        """No-match row should show '—' in file path column."""
        result, orphan, output_path = sample_results
        no_match = MatchResult(
            ledger_row=result.ledger_row,
            matched_file=None,
            score=0.0,
            status="No Match",
        )
        write_results([no_match], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        assert ws.cell(row=2, column=7).value == "—"


class TestAlternativesColumn:
    def test_alternatives_shown(self, sample_results):
        """Alternatives column should list runner-up candidates."""
        result, orphan, output_path = sample_results
        write_results([result], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        alt_val = ws.cell(row=2, column=13).value
        assert "alt_file.pdf" in alt_val
        assert "score=75" in alt_val

    def test_no_alternatives_empty(self, sample_results):
        """Row with no alternatives should have empty alternatives cell."""
        result, orphan, output_path = sample_results
        result.alternatives = []
        write_results([result], [orphan], output_path)
        wb = load_workbook(output_path)
        ws = wb["Match Results"]
        alt_val = ws.cell(row=2, column=13).value
        assert alt_val in ("", None)
